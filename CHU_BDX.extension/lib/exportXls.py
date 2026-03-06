import clr
import re
import os
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from Autodesk.Revit.DB import *
from RevitServices.Persistence import DocumentManager

clr.AddReference('RevitAPI')
clr.AddReference('RevitServices')
clr.AddReference('System.Windows.Forms')
clr.AddReference("System.IO")
from System.Windows.Forms import SaveFileDialog, DialogResult, MessageBox, MessageBoxButtons, MessageBoxIcon
from System.IO import File

# Obtenir le document actif dans Revit
doc = DocumentManager.Instance.CurrentDBDocument

# Création du dialogue de sauvegarde pour un fichier Excel
dialog = SaveFileDialog()
dialog.Filter = "Fichiers Excel (*.xls, *.xlsx)|*.xls;*.xlsx"

# Affichage de la boîte de dialogue
if dialog.ShowDialog() == DialogResult.OK:
    filename = dialog.FileName
    # Vérification de l'existence du fichier
    if not os.path.exists(filename):
        # Création d'un fichier vide (attention : ce fichier ne sera pas un classeur Excel valide,
        # il est juste "créé" pour exister sur le disque)
        fs = File.Create(filename)
        fs.Close()
    file_path = filename
else:
    file_path = None

# Fonction pour collecter et trier les éléments par catégories
def collect_and_sort_elements(doc):
    collector = FilteredElementCollector(doc).WherePasses(
        LogicalOrFilter(ElementIsElementTypeFilter(False), ElementIsElementTypeFilter(True))
    )
    sorted_elements = defaultdict(list)
    for element in collector:
        category = element.Category
        if category and category.Name:
            sorted_elements[category.Name].append(element)
    return sorted_elements

# Fonction pour analyser les paramètres des éléments
def analyze_parameters(sorted_elements):
    results = {}
    for category_name, elements in sorted_elements.items():
        parameter_names = {param.Definition.Name for element in elements if hasattr(element, "Parameters")
                           for param in element.Parameters if param.Definition and param.Definition.Name}
        parameter_names = sorted(parameter_names)
        
        category_result = []
        for element in elements:
            element_data = {
                "Element Unique ID": element.UniqueId,
                "Element ID": element.Id.IntegerValue,
                "IsType": isinstance(element, ElementType)
            }
            for param_name in parameter_names:
                param = element.LookupParameter(param_name)
                try:
                    element_data[param_name] = param.AsValueString() if param else "N/A"
                except:
                    element_data[param_name] = "N/A"
            category_result.append(element_data)
        
        results[category_name] = {
            "Parameter Names": parameter_names,
            "Elements": category_result
        }
    return results

def clean_sheet_name(name):
    """
    Nettoie le nom d'une feuille pour Excel :
      - remplace tous les caractères interdits par '_'
      - restreint à 31 caractères
      - si le résultat est vide, renomme en 'Sheet'
    """
    # On ajoute \[\] pour retirer les crochets, et ' pour l'apostrophe
    invalid_chars = r'[\/:*?"<>|\[\]\']'
    cleaned = re.sub(invalid_chars, '_', name)
    cleaned = cleaned.strip()     # supprime espaces en début/fin
    if not cleaned:
        cleaned = 'Sheet'
    return cleaned[:31]

# Fonction pour appliquer un style aux en-têtes Excel
def style_headers(ws, headers):
    title_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
    bold_font = Font(bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = title_fill
        cell.font = bold_font

# Fonction pour exporter les données vers Excel avec feuilles triées par ordre alphabétique avec largeur de colonne automatique
def export_to_excel(results, file_path):
    wb = Workbook()
    wb.remove(wb.active)  # Supprime la feuille par défaut
    for category_name in sorted(results.keys()):  # Trier les noms de catégories par ordre alphabétique
        data = results[category_name]
        ws = wb.create_sheet(title=clean_sheet_name(category_name))
        headers = ["Element Unique ID", "Element ID", "IsType"] + list(data["Parameter Names"])
        style_headers(ws, headers)
        for element in data["Elements"]: # Ajouter les lignes de données
            row = [element["Element Unique ID"], element["Element ID"], element["IsType"]] + \
                  [element.get(param_name, "N/A") for param_name in data["Parameter Names"]]
            ws.append(row)
        # Ajustement automatique de la largeur des colonnes
        for col_index, column_cells in enumerate(ws.columns, start=1):  # start=1 pour commencer à 'A'
            max_length = 0
            for cell in column_cells:
                try:
                    if cell.value:  # Si la cellule a une valeur
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Ajouter un peu d'espace supplémentaire
            ws.column_dimensions[get_column_letter(col_index)].width = adjusted_width
    wb.save(file_path)

# === Programme principal ===

# Chemin de fichier (entrée utilisateur via Dynamo)
#file_path = IN[0] if IN[0] else r"C:\\Users\\VotreNom\\Desktop\\Export_Default.xlsx"
#if not file_path.endswith(".xlsx"):
    #raise ValueError("Veuillez fournir un chemin valide se terminant par '.xlsx'.")

# Si aucun fichier n’a été sélectionné, on annule proprement
if not file_path:
    MessageBox.Show("Export annulé par l'utilisateur.", "Annulé", MessageBoxButtons.OK, MessageBoxIcon.Warning)
    OUT = "Export annulé par l'utilisateur"
    sys.exit()

# Collecte et tri des éléments
sorted_elements = collect_and_sort_elements(doc)

# Analyse des paramètres
results = analyze_parameters(sorted_elements)

# Exportation vers Excel
export_to_excel(results, file_path)

# Résultat pour Dynamo
resultat = f"Fichier Excel exporté : {file_path}"

# Affichage de la fenêtre de notification
MessageBox.Show(resultat, "Notification", MessageBoxButtons.OK, MessageBoxIcon.Information)

OUT = resultat