# ! python3
import clr
clr.AddReference('RevitAPI')
from Autodesk.Revit.DB import *
from Autodesk.Revit.DB.Architecture import Room
from Autodesk.Revit.DB.Mechanical import Space


import os
import re
import sys


from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import PatternFill, Font


from pyrevit import revit
from System.Windows.Forms import SaveFileDialog, DialogResult, MessageBox, MessageBoxButtons, MessageBoxIcon
from System.IO import File


doc = __revit__.ActiveUIDocument.Document



# ========== FONCTION: Dialogue de sauvegarde ==========
def get_save_path(document):
    """Affiche le dialogue de sauvegarde avec le nom du document Revit par défaut."""
    dialog = SaveFileDialog()
    dialog.Filter = "Fichiers Excel (*.xlsx)|*.xlsx"
    
    revit_doc_name = document.Title if document.Title else "Export_Revit"
    invalid_chars = r'[<>:"/\\|?*]'
    clean_name = re.sub(invalid_chars, '_', revit_doc_name)
    
    if clean_name.lower().endswith('.rvt'):
        clean_name = clean_name[:-4]
    
    dialog.FileName = "{}.xlsx".format(clean_name)
    
    if dialog.ShowDialog() == DialogResult.OK:
        path = dialog.FileName
        if not os.path.exists(path):
            File.Create(path).Close()
        return path
    return None



# ========== FONCTION: Obtenir la première phase (Existant) ==========
def get_existing_phase(document):
    """
    Retourne la première phase du document (phase existante/état des lieux).
    Les phases sont toujours ordonnées de la plus ancienne à la plus récente.
    """
    phases = document.Phases
    if phases.Size > 0:
        return phases[0]
    return None



# ========== FONCTION: Vérifier les Rooms dans document ==========
def has_rooms_in_document(document):
    """Vérifie si le document contient des Rooms."""
    try:
        count = FilteredElementCollector(document) \
            .OfCategory(BuiltInCategory.OST_Rooms) \
            .WhereElementIsNotElementType() \
            .GetElementCount()
        return count > 0
    except:
        return False



# ========== FONCTION: Vérifier les Spaces dans document ==========
def has_spaces_in_document(document):
    """Vérifie si le document contient des Spaces."""
    try:
        count = FilteredElementCollector(document) \
            .OfCategory(BuiltInCategory.OST_MEPSpaces) \
            .WhereElementIsNotElementType() \
            .GetElementCount()
        return count > 0
    except:
        return False



# ========== FONCTION: Obtenir le document lié ARC avec Rooms ==========
def get_arc_linked_document(document):
    """
    Récupère le document lié ARC contenant des Rooms.
    Optimisé pour une seule maquette ARC.
    """
    try:
        link_instances = FilteredElementCollector(document) \
            .OfClass(RevitLinkInstance) \
            .ToElements()
        
        for link_instance in link_instances:
            try:
                linked_doc = link_instance.GetLinkDocument()
                
                if linked_doc and not linked_doc.IsFamilyDocument:
                    # Vérifier si c'est une maquette ARC
                    doc_name = linked_doc.Title.upper()
                    if 'ARC' in doc_name or 'ARCHI' in doc_name:
                        # Vérifier qu'elle contient des Rooms
                        if has_rooms_in_document(linked_doc):
                            return {
                                'document': linked_doc,
                                'instance': link_instance,
                                'transform': link_instance.GetTotalTransform(),
                                'name': linked_doc.Title
                            }
            except:
                continue
        
    except Exception as e:
        print("Erreur recuperation lien ARC: {}".format(str(e)))
    
    return None



# ========== FONCTION: Obtenir Room dans document actuel ==========
def get_room_in_current_doc(doc, point, phase):
    """Cherche un Room à un point dans le document actuel."""
    try:
        if phase:
            room = doc.GetRoomAtPoint(point, phase)
            if room:
                room_number = room.get_Parameter(BuiltInParameter.ROOM_NUMBER).AsString()
                return room_number if room_number else ""
    except:
        pass
    return None



# ========== FONCTION: Obtenir Room dans document lié ARC ==========
def get_room_in_linked_arc(linked_arc, point):
    """Cherche un Room dans le document lié ARC."""
    if not linked_arc:
        return None
    
    try:
        linked_doc = linked_arc['document']
        transform = linked_arc['transform']
        
        # Transformer le point du document hôte vers le document lié
        transformed_point = transform.Inverse.OfPoint(point)
        
        # Utiliser la première phase du document lié
        linked_phase = get_existing_phase(linked_doc)
        
        if linked_phase:
            room = linked_doc.GetRoomAtPoint(transformed_point, linked_phase)
            if room:
                room_number = room.get_Parameter(BuiltInParameter.ROOM_NUMBER).AsString()
                return room_number if room_number else ""
    except Exception as e:
        print("Erreur lien ARC: {}".format(str(e)))
    
    return None



# ========== FONCTION: Obtenir Space dans document actuel ==========
def get_space_in_current_doc(doc, point, phase):
    """Cherche un Space à un point dans le document actuel."""
    try:
        if phase:
            space = doc.GetSpaceAtPoint(point, phase)
            if space:
                space_number = space.get_Parameter(BuiltInParameter.ROOM_NUMBER).AsString()
                return space_number if space_number else ""
    except:
        pass
    return None



# ========== FONCTION: Extrait le point de calcul spatial optimal ==========
def get_spatial_element_point(element):
    """
    Extrait le meilleur point spatial pour localiser un élément.
    Priorité: Point de calcul spatial (vert) > Point d'origine > Centre de courbe
    """
    try:
        # Vérifier si l'élément est une FamilyInstance avec point de calcul
        if isinstance(element, FamilyInstance):
            try:
                if element.HasSpatialElementCalculationPoint:
                    return element.GetSpatialElementCalculationPoint()
            except:
                pass
    except:
        pass
    
    # Fallback: utiliser Location standard
    location = element.Location
    if location is None:
        return None
    
    try:
        if isinstance(location, LocationPoint):
            return location.Point
        elif isinstance(location, LocationCurve):
            curve = location.Curve
            return curve.Evaluate(0.5, True)
    except:
        pass
    
    return None



# ========== FONCTION: Obtenir numéro de Room ou Space OPTIMISÉE ==========
def get_spatial_number(doc, element, phase, has_rooms, 
                                 has_spaces, linked_arc):
    """
    Stratégie optimisée avec GetSpatialElementCalculationPoint():
    1. Point de calcul spatial si disponible (FamilyInstance)
    2. Point d'insertion standard (LocationPoint/LocationCurve)
    3. Room dans document actuel
    4. Room dans maquette ARC liée
    5. Space dans document actuel
    """
    try:
        # Obtenir le point optimal (calcul spatial en priorité)
        point = get_spatial_element_point(element)
        
        if point is None:
            return ""
        
        # 1. Chercher Room dans document actuel
        if has_rooms:
            room_number = get_room_in_current_doc(doc, point, phase)
            if room_number:
                return room_number
        
        # 2. Chercher Room dans maquette ARC liée
        if linked_arc:
            room_number = get_room_in_linked_arc(linked_arc, point)
            if room_number:
                return room_number
        
        # 3. Chercher Space dans document actuel
        if has_spaces:
            space_number = get_space_in_current_doc(doc, point, phase)
            if space_number:
                return space_number
        
    except Exception as e:
        print("Erreur detection spatiale optimisee: {}".format(str(e)))
    
    return ""



# ========== FONCTION: Collecte optimisée par catégorie ==========
def collect_and_sort_elements(document):
    """Collecte les éléments en filtrant directement par catégorie."""
    sorted_elements = {}
    
    # Collecter tous les instances en un seul appel
    all_instances = FilteredElementCollector(document) \
        .WhereElementIsNotElementType() \
        .ToElements()
    
    # Collecter tous les types en un seul appel
    all_types = FilteredElementCollector(document) \
        .WhereElementIsElementType() \
        .ToElements()
    
    # Regrouper par catégorie (une seule boucle)
    for element in list(all_instances) + list(all_types):
        if element.Category is not None:  # Filtrer les None
            cat_name = element.Category.Name
            if cat_name not in sorted_elements:
                sorted_elements[cat_name] = []
            sorted_elements[cat_name].append(element)
    
    return sorted_elements


# ========== FONCTION: Nettoyage des noms de feuilles ==========
def clean_sheet_name(name, existing_names):
    """Nettoie et valide les noms de feuilles Excel (max 31 caractères)."""
    max_length = 31
    invalid_chars = r'[\\/*?:\[\]]'
    base = re.sub(invalid_chars, '_', name)[:max_length]
    
    if base not in existing_names:
        return base
    
    for i in range(1, 1000):
        suffix = f"_{i}"
        new_name = f"{base[:max_length - len(suffix)]}{suffix}"
        if new_name not in existing_names:
            return new_name
    return base[:max_length]



# ========== FONCTION: Création des cellules stylisées ==========
def create_styled_header_cells(ws, headers):
    """Crée une ligne d'en-tête stylisée pour le mode write_only."""
    header_cells = []
    fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')
    font = Font(bold=True)
    
    for header_text in headers:
        cell = WriteOnlyCell(ws, value=header_text)
        cell.fill = fill
        cell.font = font
        header_cells.append(cell)
    
    return header_cells



# ========== FONCTION: Export optimisé ==========
def export_to_excel_optimized(sorted_elements, path, document, phase, 
                              has_rooms, has_spaces, linked_arc):
    """Export avec détection Room/Space (actuel + lien ARC) et point spatial optimisé."""
    wb = Workbook(write_only=True)
    used_sheet_names = set()
    
    for cat_name, elements in sorted(sorted_elements.items()):
        try:
            param_names = set()
            for el in elements:
                for param in el.Parameters:
                    if param.Definition:
                        param_names.add(param.Definition.Name)
            param_names = sorted(param_names)
            
            sheet_name = clean_sheet_name(cat_name, used_sheet_names)
            used_sheet_names.add(sheet_name)
            ws = wb.create_sheet(title=sheet_name)
            
            headers = ['Element Unique ID', 'Element ID', 'Type/Occurrence', 
                      'RGU_Number'] + param_names
            styled_headers = create_styled_header_cells(ws, headers)
            ws.append(styled_headers)
            
            for el in elements:
                spatial_number = get_spatial_number(
                    document, el, phase, has_rooms, has_spaces, linked_arc
                )
                
                is_type = isinstance(el, ElementType)

                row = [
                    el.UniqueId,
                    el.Id.IntegerValue,
                    "Type" if is_type else "Occurrence",
                    spatial_number
                ]
                
                for name in param_names:
                    param = el.LookupParameter(name)
                    value = param.AsValueString() if param and param.HasValue else ""
                    row.append(value if value else "")
                
                ws.append(row)
        
        except Exception as e:
            print("Erreur catégorie {}: {}".format(cat_name, str(e)))
            continue
    
    wb.save(path)



# ========== FONCTION: Export avec gestion d'erreurs ==========
def export_with_error_handling(document, file_path):
    """Fonction principale optimisée pour une seule maquette ARC."""
    try:
        # Obtenir la première phase (existant)
        phase = get_existing_phase(document)
        
        if not phase:
            MessageBox.Show(
                "Aucune phase trouvée dans le document.",
                "Erreur",
                MessageBoxButtons.OK,
                MessageBoxIcon.Error
            )
            return False
        
        print("Phase utilisee: {} (premiere phase)".format(phase.Name))
        
        # Vérifier document actuel
        has_rooms = has_rooms_in_document(document)
        has_spaces = has_spaces_in_document(document)
        
        # Récupérer le document lié ARC si pas de Rooms localement
        linked_arc = None
        if not has_rooms:
            print("Aucun Room dans le document actuel, recherche du lien ARC...")
            linked_arc = get_arc_linked_document(document)
            
            if linked_arc:
                print("Maquette ARC trouvee: {}".format(linked_arc['name']))
        
        # Informations pour l'utilisateur
        spatial_info = []
        if has_rooms:
            spatial_info.append("Rooms (document actuel)")
        if linked_arc:
            spatial_info.append("Rooms (lien ARC: {})".format(linked_arc['name']))
        if has_spaces:
            spatial_info.append("Spaces (document actuel)")
        
        if not has_rooms and not linked_arc and not has_spaces:
            MessageBox.Show(
                "Aucun Room ni Space trouvé dans le document actuel ou lié.",
                "Avertissement",
                MessageBoxButtons.OK,
                MessageBoxIcon.Warning
            )
        
        # Collecte et export
        sorted_elements = collect_and_sort_elements(document)
        
        if not sorted_elements:
            MessageBox.Show(
                "Aucun élément trouvé dans le projet.",
                "Avertissement",
                MessageBoxButtons.OK,
                MessageBoxIcon.Warning
            )
            return False
        
        export_to_excel_optimized(sorted_elements, file_path, document, phase,
                                  has_rooms, has_spaces, linked_arc)
        
        MessageBox.Show(
            "Export Excel terminé :\n{}\n\nPhase: {} (première phase)\nSources spatiales:\n{}".format(
                file_path,
                phase.Name,
                "\n".join("- " + info for info in spatial_info) if spatial_info else "- Aucune"
            ),
            "Succès",
            MessageBoxButtons.OK,
            MessageBoxIcon.Information
        )
        return True
    
    except Exception as e:
        MessageBox.Show(
            "Erreur lors de l'export :\n{}".format(str(e)),
            "Erreur",
            MessageBoxButtons.OK,
            MessageBoxIcon.Error
        )
        return False



# ========== PROGRAMME PRINCIPAL ==========
file_path = get_save_path(doc)


if not file_path:
    MessageBox.Show(
        "Export annulé par l'utilisateur.",
        "Annulé",
        MessageBoxButtons.OK,
        MessageBoxIcon.Warning
    )
    sys.exit()


export_with_error_handling(doc, file_path)
